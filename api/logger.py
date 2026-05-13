import os
import json
from datetime import datetime, timezone, timedelta
from openpyxl import Workbook, load_workbook
from google.oauth2 import service_account
from googleapiclient.discovery import build

# Cấu hình đường dẫn lưu log
LOG_DIR = 'logs'
LOG_FILE = os.path.join(LOG_DIR, 'analysis_history.xlsx')
SERVICE_ACCOUNT_FILE = 'secrets/gemini-account.json'
GOOGLE_SHEET_ID = os.environ.get('GOOGLE_SHEET_ID')

def ensure_log_exists():
    """Đảm bảo thư mục và tệp Excel tồn tại."""
    if not os.path.exists(LOG_DIR):
        os.makedirs(LOG_DIR)
    
    if not os.path.exists(LOG_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Analysis History"
        # Định nghĩa tiêu đề cột
        headers = ["Thời gian", "Nội dung tin nhắn", "Nguy hiểm?", "Loại hình", "Lý do", "Mức độ", "Khuyến cáo"]
        ws.append(headers)
        wb.save(LOG_FILE)
        print(f"📁 [Logger] Đã tạo tệp log mới tại: {LOG_FILE}")

def log_to_excel(row_data):
    """Lưu vào file Excel nội bộ (Backup)."""
    try:
        ensure_log_exists()
        wb = load_workbook(LOG_FILE)
        ws = wb.active
        ws.append(row_data)
        wb.save(LOG_FILE)
        print(f"✅ [Logger] Đã ghi log vào Excel.")
    except Exception as e:
        print(f"🔴 [Logger] Lỗi Excel: {e}")

def log_to_google_sheets(row_data):
    """Lưu vào Google Sheets sử dụng Service Account."""
    if not GOOGLE_SHEET_ID:
        print("🟡 [Logger] GOOGLE_SHEET_ID chưa được thiết lập. Bỏ qua lưu Sheets.")
        return

    if not os.path.exists(SERVICE_ACCOUNT_FILE):
        print(f"🟡 [Logger] Không tìm thấy {SERVICE_ACCOUNT_FILE}. Bỏ qua lưu Sheets.")
        return

    try:
        creds = service_account.Credentials.from_service_account_file(
            SERVICE_ACCOUNT_FILE, 
            scopes=['https://www.googleapis.com/auth/spreadsheets']
        )
        # Tắt cache để tránh cảnh báo file_cache
        service = build('sheets', 'v4', credentials=creds, cache_discovery=False)
        
        body = {'values': [row_data]}
        # Ghi vào sheet có tên là 'History'. Đảm bảo tab trong Google Sheet của bạn được đặt tên là History.
        service.spreadsheets().values().append(
            spreadsheetId=GOOGLE_SHEET_ID,
            range='History!A2', 
            valueInputOption='USER_ENTERED',
            insertDataOption='INSERT_ROWS',
            body=body
        ).execute()
        print("✅ [Logger] Đã ghi log vào Google Sheets thành công.")
    except Exception as e:
        print(f"🔴 [Logger] Lỗi Google Sheets: {e}")

# --- Audit Log cho Admin ---
ADMIN_LOG_FILE = os.path.join(LOG_DIR, 'admin_actions.log')

def audit_log(action: str, admin_name: str = "Unknown"):
    """Ghi lại các hoạt động của admin vào file log riêng."""
    try:
        if not os.path.exists(LOG_DIR):
            os.makedirs(LOG_DIR)
        
        vn_tz = timezone(timedelta(hours=7))
        timestamp = datetime.now(vn_tz).strftime('%Y-%m-%d %H:%M:%S')
        log_entry = f"[{timestamp}] Admin: {admin_name} | Action: {action}\n"
        
        with open(ADMIN_LOG_FILE, 'a', encoding='utf-8') as f:
            f.write(log_entry)
        print(f"🔒 [Audit] {log_entry.strip()}")
    except Exception as e:
        print(f"🔴 [Audit] Lỗi ghi log: {e}")

def log_analysis(text: str, result: dict):
    """Lưu kết quả phân tích vào cả Excel và Google Sheets."""
    # Lấy thời gian hiện tại (VN)
    vn_tz = timezone(timedelta(hours=7))
    timestamp = datetime.now(vn_tz).strftime('%Y-%m-%d %H:%M:%S')

    # Chuẩn bị dữ liệu hàng
    is_dangerous = result.get('is_dangerous', False)
    types = ", ".join(result.get('types', [])) if isinstance(result.get('types'), list) else str(result.get('types', 'N/A'))
    
    row = [
        timestamp,
        text[:1000],
        "CÓ" if is_dangerous else "KHÔNG",
        types,
        result.get('reason', 'N/A'),
        f"{result.get('score', 0)}/5",
        result.get('recommend', 'N/A')
    ]

    # Ghi log đồng thời (nên chạy trong background threads từ analyze.py)
    log_to_excel(row)
    log_to_google_sheets(row)
